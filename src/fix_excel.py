"""修复被测试数据污染的大表"""
import pandas as pd
from openpyxl import load_workbook

path = 'data/output/all_in主动基金风控纪律手册.xlsx'

# 加载工作簿
wb = load_workbook(path)
ws = wb['季度检查表']

# 查看当前有多少列
print(f"修复前：{ws.max_column} 列，{ws.max_row} 行")

# 删除第16列到最后一列（索引从1开始，第16列就是被污染的）
# 从右往左删，避免索引偏移
for col in range(ws.max_column, 15, -1):
    ws.delete_cols(col)

# 检查第2行（header行）是否有测试数据残留在右边
# 现在只保留15列，再清理第2行第16列以后的内容（已经删了）

print(f"修复后：{ws.max_column} 列，{ws.max_row} 行")

# 同时检查第1行的标题是否正确
print(f"第1行A1: {ws.cell(1,1).value}")
print(f"第2行列名: {[ws.cell(2, c).value for c in range(1, ws.max_column+1)]}")

wb.save(path)
wb.close()
print("\n修复完成！")