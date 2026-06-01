"""
Excel 读写引擎 v2
正确处理大表格式：第1行标题说明，第2行列名，第3行起数据
"""

import json
import os
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook


# ============ 配置加载 ============

def load_config():
    """加载基金配置"""
    config_path = os.path.join(os.path.dirname(__file__), '..', 'config', 'fund_config.json')
    with open(config_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def get_master_excel_path():
    """获取大表完整路径"""
    config = load_config()
    base_dir = os.path.join(os.path.dirname(__file__), '..')
    return os.path.join(base_dir, config['master_excel_path'])


# ============ 读取大表 ============

def read_sheet(sheet_name):
    """
    读取大表中指定 sheet，跳过第1行标题，用第2行作为列名。
    返回 DataFrame（数据从第3行开始）
    """
    path = get_master_excel_path()
    if not os.path.exists(path):
        raise FileNotFoundError(f"大表不存在: {path}")
    return pd.read_excel(path, sheet_name=sheet_name, header=1)


def get_sheet_names():
    """获取大表所有 sheet 名称"""
    path = get_master_excel_path()
    wb = load_workbook(path, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def get_sheet_columns(sheet_name):
    """获取指定 sheet 的列名列表"""
    path = get_master_excel_path()
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet_name]
    # 第2行是列名
    columns = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    columns = [c for c in columns if c is not None]
    wb.close()
    return columns


# ============ upsert 回写 ============

def upsert_to_sheet(sheet_name, new_data, key_columns):
    """
    upsert 逻辑：按 key_columns 匹配，存在则更新该行，不存在则追加到末尾。
    保留第1行标题和第2行列名不动，只操作第3行以下的数据区域。

    参数：
        sheet_name: 目标 sheet 名称
        new_data: dict，一行数据，key 必须匹配 sheet 的列名
        key_columns: list，用于匹配的列名，如 ["报告期"]
    """
    path = get_master_excel_path()
    wb = load_workbook(path)
    ws = wb[sheet_name]

    # 读取第2行列名
    max_col = ws.max_column
    columns = [ws.cell(2, c).value for c in range(1, max_col + 1)]

    # 找到 key_columns 对应的列索引
    key_col_indices = {}
    for kc in key_columns:
        if kc in columns:
            key_col_indices[kc] = columns.index(kc) + 1  # openpyxl 从1开始
        else:
            print(f"[警告] 列 '{kc}' 不存在于 sheet '{sheet_name}'")
            wb.close()
            return

    # 搜索数据区域（第3行开始）找匹配行
    match_row = None
    for row in range(3, ws.max_row + 1):
        all_match = True
        for kc, col_idx in key_col_indices.items():
            cell_val = ws.cell(row, col_idx).value
            if str(cell_val) != str(new_data.get(kc, '')):
                all_match = False
                break
        if all_match:
            match_row = row
            break

    if match_row:
        # 更新已有行
        for col_name, value in new_data.items():
            if col_name in columns:
                col_idx = columns.index(col_name) + 1
                ws.cell(match_row, col_idx, value)
        print(f"[更新] sheet '{sheet_name}' 第{match_row}行 | {[new_data.get(k) for k in key_columns]}")
    else:
        # 追加新行
        new_row_num = ws.max_row + 1
        for col_name, value in new_data.items():
            if col_name in columns:
                col_idx = columns.index(col_name) + 1
                ws.cell(new_row_num, col_idx, value)
        print(f"[新增] sheet '{sheet_name}' 第{new_row_num}行 | {[new_data.get(k) for k in key_columns]}")

    wb.save(path)
    wb.close()


def append_to_sheet(sheet_name, new_data):
    """
    纯追加模式（用于执行记录等不需要去重的表）
    """
    path = get_master_excel_path()
    wb = load_workbook(path)
    ws = wb[sheet_name]

    # 读取第2行列名
    max_col = ws.max_column
    columns = [ws.cell(2, c).value for c in range(1, max_col + 1)]

    # 追加到末尾
    new_row_num = ws.max_row + 1
    for col_name, value in new_data.items():
        if col_name in columns:
            col_idx = columns.index(col_name) + 1
            ws.cell(new_row_num, col_idx, value)

    wb.save(path)
    wb.close()
    print(f"[追加] sheet '{sheet_name}' 第{new_row_num}行")


# ============ 测试入口 ============

if __name__ == '__main__':
    print("=== Excel 引擎 v2 测试 ===")
    print(f"大表路径: {get_master_excel_path()}")
    print(f"现有 sheets: {get_sheet_names()}")

    # 显示季度检查表的列名
    cols = get_sheet_columns('季度检查表')
    print(f"\n季度检查表列名: {cols}")

    # 测试 upsert：写入一条测试数据
    test_data = {
        "检查日期": "2025-07-01",
        "报告期": "2025Q2_TEST",
        "李晓星在任?": "是",
        "张萍在任?": "是",
        "杜宇在任?": "是",
        "基金规模(亿元)": 45,
        "6个月规模是否翻倍?": "否",
        "股票仓位": 85,
        "前十大集中度": 48,
        "防御资产占比": 15,
        "观察触发数": 0,
        "本期结论": "绿灯-测试"
    }

    upsert_to_sheet("季度检查表", test_data, ["报告期"])
    print("\n再次写入相同报告期（应该更新而非新增）...")
    test_data["本期结论"] = "绿灯-测试更新"
    upsert_to_sheet("季度检查表", test_data, ["报告期"])

    print("\n测试完成！打开大表检查'季度检查表'最后几行。")